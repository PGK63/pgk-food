#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import questionary  # type: ignore
except Exception:  # noqa: BLE001
    questionary = None


STUDENT_ROLE = "STUDENT"
CURATOR_ROLE = "CURATOR"

CATEGORY_MAP = {
    "сво": "SVO",
    "многодетные": "MANY_CHILDREN",
}

REJECT_HEADER = [
    "sheet",
    "row",
    "error_code",
    "error_message",
    "raw_fio",
    "raw_group",
    "raw_category",
    "raw_curator",
]

CURATORS_SHEET_HEADER = ["ФИО", "Группы", "Логин", "Пароль", "userId", "Статус"]
STUDENTS_SHEET_HEADER = [
    "ЛистИсточник",
    "Строка",
    "ФИО",
    "Группа",
    "Категория",
    "Куратор",
    "Логин",
    "Пароль",
    "userId",
    "Статус",
]


class ImportFailure(RuntimeError):
    pass


class ApiError(RuntimeError):
    def __init__(self, message: str, status_code: Optional[int] = None, response_body: Optional[str] = None):
        super().__init__(message)
        self.status_code = status_code
        self.response_body = response_body


FioParts = Tuple[str, str, str]


@dataclass(frozen=True)
class CuratorKey:
    surname: str
    name: str
    father_name: str

    @property
    def full_name(self) -> str:
        return f"{self.surname} {self.name} {self.father_name}"


@dataclass(frozen=True)
class StudentKey:
    surname: str
    name: str
    father_name: str
    group_name: str

    @property
    def full_name(self) -> str:
        return f"{self.surname} {self.name} {self.father_name}"


@dataclass
class ParsedStudent:
    sheet: str
    row: int
    surname: str
    name: str
    father_name: str
    full_name: str
    group_name: str
    category: str
    curator: CuratorKey
    raw_category: str
    raw_group: str
    raw_curator: str
    raw_comment: str
    raw_breakfast: str
    raw_lunch: str

    @property
    def key(self) -> StudentKey:
        return StudentKey(
            surname=self.surname,
            name=self.name,
            father_name=self.father_name,
            group_name=self.group_name,
        )


@dataclass
class RejectRow:
    sheet: str
    row: int
    error_code: str
    error_message: str
    raw_fio: str
    raw_group: str
    raw_category: str
    raw_curator: str


@dataclass
class ParsedWorkbook:
    source_path: str
    processed_student_rows: int = 0
    skipped_non_student_rows: int = 0
    valid_students: List[ParsedStudent] = field(default_factory=list)
    rejects: List[RejectRow] = field(default_factory=list)
    sheets_seen: List[str] = field(default_factory=list)
    breakfast_values: Counter = field(default_factory=Counter)
    lunch_values: Counter = field(default_factory=Counter)
    comments_non_empty: int = 0

    @property
    def groups(self) -> List[str]:
        return sorted({student.group_name for student in self.valid_students})

    @property
    def curators(self) -> List[CuratorKey]:
        return sorted(
            {student.curator for student in self.valid_students},
            key=lambda x: (x.surname, x.name, x.father_name),
        )

    @property
    def group_curators(self) -> Dict[str, List[CuratorKey]]:
        mapping: Dict[str, set[CuratorKey]] = defaultdict(set)
        for student in self.valid_students:
            mapping[student.group_name].add(student.curator)
        return {
            group_name: sorted(curators, key=lambda x: (x.surname, x.name, x.father_name))
            for group_name, curators in sorted(mapping.items(), key=lambda kv: kv[0])
        }

    @property
    def stats(self) -> Dict[str, int]:
        return {
            "processed_student_rows": self.processed_student_rows,
            "skipped_non_student_rows": self.skipped_non_student_rows,
            "valid_students": len(self.valid_students),
            "rejects": len(self.rejects),
            "groups": len(self.groups),
            "curators": len(self.curators),
        }


@dataclass
class CuratorCredentialRow:
    curator: CuratorKey
    groups: List[str]
    login: str = ""
    password: str = ""
    user_id: str = ""
    status: str = "PLANNED"


@dataclass
class StudentCredentialRow:
    sheet: str
    row: int
    full_name: str
    group_name: str
    category: str
    curator_full_name: str
    login: str = ""
    password: str = ""
    user_id: str = ""
    status: str = "PLANNED"


@dataclass
class ImportState:
    created_group_ids: List[int] = field(default_factory=list)
    created_curator_ids: List[str] = field(default_factory=list)
    created_student_ids: List[str] = field(default_factory=list)
    deleted_group_ids: List[int] = field(default_factory=list)
    deleted_curator_ids: List[str] = field(default_factory=list)
    deleted_student_ids: List[str] = field(default_factory=list)


@dataclass
class ImportResult:
    success: bool
    preflight_ok: bool
    applied: bool
    error: Optional[str]
    rollback_performed: bool
    rollback_success: bool
    rollback_errors: List[str]
    group_id_map: Dict[str, int]
    curator_rows: Dict[CuratorKey, CuratorCredentialRow]
    student_rows: List[StudentCredentialRow]
    state: ImportState


@dataclass
class PreflightBaseline:
    ok: bool
    message: Optional[str]
    ignored_current_admin: bool
    current_user_id: Optional[str]
    current_user_group_id: Optional[int]
    raw_counts: Dict[str, int]
    effective_counts: Dict[str, int]


@dataclass
class InteractiveFixResult:
    enabled: bool
    total_issues: int
    fixed_rows: int
    skipped_rows: int
    aborted: bool
    fixed_xlsx_path: Optional[str]


class InteractivePrompter:
    def choose_action(self, reject: "RejectRow", current_values: Dict[str, str]) -> str:
        raise NotImplementedError

    def ask_text(self, label: str, current_value: str) -> str:
        raise NotImplementedError

    def ask_category(self, current_value: str) -> str:
        raise NotImplementedError


class TerminalPrompter(InteractivePrompter):
    def __init__(self) -> None:
        if not sys.stdin.isatty():
            raise ImportFailure("Интерактивная правка требует запуска в терминале (TTY)")

    def choose_action(self, reject: "RejectRow", current_values: Dict[str, str]) -> str:
        message = (
            f"\n[{reject.sheet}:{reject.row}] {reject.error_code}\n"
            f"{reject.error_message}\n"
            f"ФИО: {current_values.get('student_fio', '')}\n"
            f"Группа: {current_values.get('group', '')}\n"
            f"Категория: {current_values.get('category', '')}\n"
            f"Куратор: {current_values.get('curator', '')}\n"
            "Выберите действие:"
        )
        choices = [
            ("Редактировать", "edit"),
            ("Пропустить", "skip"),
            ("Прервать импорт", "abort"),
        ]
        if questionary is not None:
            answer = questionary.select(
                message,
                choices=[questionary.Choice(title=title, value=value) for title, value in choices],
            ).ask()
            if answer is None:
                return "abort"
            return str(answer)

        print(message)
        for idx, (title, value) in enumerate(choices, start=1):
            print(f"{idx}. {title} [{value}]")
        while True:
            raw = input("> ").strip().lower()
            if raw in {"1", "edit", "e"}:
                return "edit"
            if raw in {"2", "skip", "s"}:
                return "skip"
            if raw in {"3", "abort", "a", "q"}:
                return "abort"
            print("Введите 1/2/3")

    def ask_text(self, label: str, current_value: str) -> str:
        if questionary is not None:
            answer = questionary.text(label, default=current_value).ask()
            return current_value if answer is None else str(answer)
        raw = input(f"{label} [{current_value}]: ").strip()
        return current_value if raw == "" else raw

    def ask_category(self, current_value: str) -> str:
        normalized = normalize_category_label(current_value)
        choices = ["СВО", "Многодетные"]
        default_choice = normalized if normalized in choices else "Многодетные"
        if questionary is not None:
            answer = questionary.select("Категория", choices=choices, default=default_choice).ask()
            return default_choice if answer is None else str(answer)
        raw = input(f"Категория [СВО/Многодетные] [{default_choice}]: ").strip()
        if not raw:
            return default_choice
        return normalize_category_label(raw) or raw


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_key(value: Any) -> str:
    return normalize_space(value).lower().replace("ё", "е")


def split_camel_token(token: str) -> List[str]:
    if not token:
        return []
    parts: List[str] = []
    start = 0
    for idx in range(1, len(token)):
        prev = token[idx - 1]
        cur = token[idx]
        if cur.isupper() and prev.islower():
            parts.append(token[start:idx])
            start = idx
    parts.append(token[start:])
    return [part for part in parts if part]


def split_fio(raw_fio: str, allow_camel_fix: bool) -> Tuple[Optional[FioParts], str]:
    fio = normalize_space(raw_fio)
    if not fio:
        return None, "empty"

    tokens = fio.split(" ")
    if len(tokens) == 3:
        return (tokens[0], tokens[1], tokens[2]), "as_is"

    if not allow_camel_fix:
        return None, f"invalid_parts_{len(tokens)}"

    expanded: List[str] = []
    for token in tokens:
        expanded.extend(split_camel_token(token))

    if len(expanded) == 3:
        return (expanded[0], expanded[1], expanded[2]), "camel_case_fixed"

    return None, f"invalid_parts_{len(expanded)}"


def normalize_category_label(value: str) -> str:
    key = normalize_key(value).replace(".", "")
    if key == "сво":
        return "СВО"
    if key == "многодетные":
        return "Многодетные"
    return normalize_space(value)


def parse_bool_mark(value: str) -> bool:
    v = normalize_key(value)
    return v in {"1", "да", "true", "x", "х"}


def map_category(raw_category: str) -> Optional[str]:
    raw = normalize_key(raw_category).replace(".", "")
    return CATEGORY_MAP.get(raw)


def parse_header(row_values: Sequence[Any]) -> Dict[str, int]:
    header: Dict[str, int] = {}

    for idx, value in enumerate(row_values, start=1):
        key = normalize_key(value)
        if not key:
            continue

        if key == "фио обучающегося":
            header["student_fio"] = idx
        elif key == "статус":
            header["status"] = idx
        elif key == "группа":
            header["group"] = idx
        elif key == "категория":
            header["category"] = idx
        elif key == "завтрак":
            header["breakfast"] = idx
        elif key == "обед":
            header["lunch"] = idx
        elif key.startswith("комментарии"):
            header["comment"] = idx
        elif key == "классный руководитель":
            header["curator"] = idx

    required = ["student_fio", "status", "group", "category", "curator"]
    missing = [field for field in required if field not in header]
    if missing:
        raise ImportFailure(f"Не найдены обязательные колонки в заголовке: {', '.join(missing)}")

    return header


def read_row_values_from_sheet(ws: Worksheet, row_idx: int, header: Dict[str, int]) -> Dict[str, str]:
    def col(name: str) -> str:
        idx = header.get(name)
        return normalize_space(ws.cell(row=row_idx, column=idx).value if idx else "")

    return {
        "student_fio": col("student_fio"),
        "status": col("status"),
        "group": col("group"),
        "category": col("category"),
        "breakfast": col("breakfast"),
        "lunch": col("lunch"),
        "comment": col("comment"),
        "curator": col("curator"),
    }


def validate_row(
    *,
    sheet: str,
    row: int,
    row_values: Dict[str, str],
    seen_students: Optional[set[StudentKey]] = None,
    check_duplicates: bool = True,
) -> Tuple[Optional[ParsedStudent], List[Tuple[str, str]]]:
    raw_student_fio = row_values.get("student_fio", "")
    raw_group = row_values.get("group", "")
    raw_category = row_values.get("category", "")
    raw_breakfast = row_values.get("breakfast", "")
    raw_lunch = row_values.get("lunch", "")
    raw_comment = row_values.get("comment", "")
    raw_curator = row_values.get("curator", "")

    student_parts, student_mode = split_fio(raw_student_fio, allow_camel_fix=True)
    curator_parts, curator_mode = split_fio(raw_curator, allow_camel_fix=False)
    category = map_category(raw_category)

    errors: List[Tuple[str, str]] = []
    if student_parts is None:
        errors.append(("bad_student_fio", f"Невалидное ФИО студента ({student_mode})"))
    if not raw_group:
        errors.append(("empty_group", "Не указана группа"))
    if category is None:
        errors.append(("bad_category", f"Неизвестная категория: '{raw_category}'"))
    if curator_parts is None:
        errors.append(("bad_curator_fio", f"Невалидное ФИО куратора ({curator_mode})"))

    if not errors and student_parts is not None and check_duplicates and seen_students is not None:
        student_key = StudentKey(
            surname=student_parts[0],
            name=student_parts[1],
            father_name=student_parts[2],
            group_name=raw_group,
        )
        if student_key in seen_students:
            errors.append(("duplicate_student", "Дубликат студента в одной группе"))
        else:
            seen_students.add(student_key)

    if errors:
        return None, errors

    assert student_parts is not None
    assert curator_parts is not None
    assert category is not None
    parsed_student = ParsedStudent(
        sheet=sheet,
        row=row,
        surname=student_parts[0],
        name=student_parts[1],
        father_name=student_parts[2],
        full_name=f"{student_parts[0]} {student_parts[1]} {student_parts[2]}",
        group_name=raw_group,
        category=category,
        curator=CuratorKey(
            surname=curator_parts[0],
            name=curator_parts[1],
            father_name=curator_parts[2],
        ),
        raw_category=raw_category,
        raw_group=raw_group,
        raw_curator=raw_curator,
        raw_comment=raw_comment,
        raw_breakfast=raw_breakfast,
        raw_lunch=raw_lunch,
    )
    return parsed_student, []


def parse_workbook(xlsx_path: Path) -> ParsedWorkbook:
    if not xlsx_path.exists():
        raise ImportFailure(f"Файл не найден: {xlsx_path}")

    wb = load_workbook(filename=xlsx_path, data_only=True)
    parsed = ParsedWorkbook(source_path=str(xlsx_path))
    seen_students: set[StudentKey] = set()

    for ws in wb.worksheets:
        parsed.sheets_seen.append(ws.title)
        if ws.max_row < 2:
            continue

        header = parse_header([cell.value for cell in ws[1]])

        for row_idx in range(2, ws.max_row + 1):
            row_values = read_row_values_from_sheet(ws, row_idx, header)
            raw_student_fio = row_values["student_fio"]
            if not raw_student_fio:
                continue

            raw_status = row_values["status"]
            if normalize_key(raw_status) != "студент":
                parsed.skipped_non_student_rows += 1
                continue

            parsed.processed_student_rows += 1

            raw_group = row_values["group"]
            raw_category = row_values["category"]
            raw_breakfast = row_values["breakfast"]
            raw_lunch = row_values["lunch"]
            raw_comment = row_values["comment"]
            raw_curator = row_values["curator"]

            parsed.breakfast_values[normalize_key(raw_breakfast) or "<empty>"] += 1
            parsed.lunch_values[normalize_key(raw_lunch) or "<empty>"] += 1
            if raw_comment:
                parsed.comments_non_empty += 1

            parsed_student, errors = validate_row(
                sheet=ws.title,
                row=row_idx,
                row_values=row_values,
                seen_students=seen_students,
                check_duplicates=True,
            )
            if errors:
                parsed.rejects.append(
                    RejectRow(
                        sheet=ws.title,
                        row=row_idx,
                        error_code=";".join(code for code, _ in errors),
                        error_message="; ".join(msg for _, msg in errors),
                        raw_fio=raw_student_fio,
                        raw_group=raw_group,
                        raw_category=raw_category,
                        raw_curator=raw_curator,
                    )
                )
                continue

            assert parsed_student is not None
            parsed.valid_students.append(parsed_student)

    return parsed


class ApiClient:
    def __init__(self, base_url: str, timeout_seconds: float = 15.0, max_attempts: int = 4):
        self.base_url = base_url.rstrip("/")
        self.timeout_seconds = timeout_seconds
        self.max_attempts = max(1, max_attempts)
        self.session = requests.Session()
        self.token: Optional[str] = None

    def _sleep_backoff(self, attempt: int) -> None:
        # Exponential backoff: 0.5, 1, 2 seconds...
        delay = 0.5 * (2 ** (attempt - 1))
        time.sleep(delay)

    @staticmethod
    def _response_error_message(resp: requests.Response) -> str:
        try:
            payload = resp.json()
        except ValueError:
            payload = None

        if isinstance(payload, dict):
            for key in ("userMessage", "message", "error"):
                value = payload.get(key)
                if isinstance(value, str) and value.strip():
                    return value
        text = (resp.text or "").strip()
        return text or f"HTTP {resp.status_code}"

    def _request(
        self,
        method: str,
        path: str,
        *,
        expected_statuses: Iterable[int],
        params: Optional[Dict[str, Any]] = None,
        json_body: Optional[Dict[str, Any]] = None,
    ) -> requests.Response:
        url = f"{self.base_url}{path}"
        headers: Dict[str, str] = {}
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"

        expected = set(expected_statuses)
        last_network_error: Optional[Exception] = None

        for attempt in range(1, self.max_attempts + 1):
            try:
                resp = self.session.request(
                    method=method,
                    url=url,
                    params=params,
                    json=json_body,
                    headers=headers,
                    timeout=self.timeout_seconds,
                )
            except requests.RequestException as exc:
                last_network_error = exc
                if attempt == self.max_attempts:
                    break
                self._sleep_backoff(attempt)
                continue

            if resp.status_code in expected:
                return resp

            if resp.status_code >= 500:
                if attempt == self.max_attempts:
                    msg = self._response_error_message(resp)
                    raise ApiError(
                        f"{method} {path} завершился 5xx после retries: {msg}",
                        status_code=resp.status_code,
                        response_body=resp.text,
                    )
                self._sleep_backoff(attempt)
                continue

            msg = self._response_error_message(resp)
            raise ApiError(
                f"{method} {path} вернул ошибку: {msg}",
                status_code=resp.status_code,
                response_body=resp.text,
            )

        raise ApiError(
            f"{method} {path} не выполнен из-за сетевой ошибки: {last_network_error}",
            status_code=None,
            response_body=str(last_network_error) if last_network_error else None,
        )

    def _request_json(
        self,
        method: str,
        path: str,
        *,
        expected_statuses: Iterable[int],
        params: Optional[Dict[str, Any]] = None,
        json_body: Optional[Dict[str, Any]] = None,
    ) -> Any:
        resp = self._request(
            method,
            path,
            expected_statuses=expected_statuses,
            params=params,
            json_body=json_body,
        )
        if not resp.content:
            return None
        try:
            return resp.json()
        except ValueError:
            return None

    def login(self, login: str, password: str) -> None:
        payload = self._request_json(
            "POST",
            "/api/v1/auth/login",
            expected_statuses=(200,),
            json_body={"login": login, "password": password},
        )
        if not isinstance(payload, dict) or not payload.get("token"):
            raise ImportFailure("Сервер не вернул токен при логине")
        self.token = payload["token"]

    def list_groups(self) -> List[Dict[str, Any]]:
        payload = self._request_json("GET", "/api/v1/groups", expected_statuses=(200,))
        if isinstance(payload, list):
            return payload
        raise ImportFailure("Невалидный ответ /api/v1/groups")

    def list_users(self, role: str) -> List[Dict[str, Any]]:
        payload = self._request_json(
            "GET",
            "/api/v1/registrator/users",
            expected_statuses=(200,),
            params={"role": role},
        )
        if isinstance(payload, list):
            return payload
        raise ImportFailure("Невалидный ответ /api/v1/registrator/users")

    def get_me(self) -> Dict[str, Any]:
        payload = self._request_json("GET", "/api/v1/auth/me", expected_statuses=(200,))
        if isinstance(payload, dict) and payload.get("userId") is not None:
            return payload
        raise ImportFailure("Невалидный ответ /api/v1/auth/me")

    def create_group(self, name: str) -> Dict[str, Any]:
        payload = self._request_json(
            "POST",
            "/api/v1/groups",
            expected_statuses=(200,),
            json_body={"name": name},
        )
        if isinstance(payload, dict) and payload.get("id") is not None:
            return payload
        raise ImportFailure("Невалидный ответ create_group")

    def create_user(
        self,
        *,
        role: str,
        name: str,
        surname: str,
        father_name: str,
        group_id: Optional[int] = None,
        student_category: Optional[str] = None,
    ) -> Dict[str, Any]:
        body: Dict[str, Any] = {
            "roles": [role],
            "name": name,
            "surname": surname,
            "fatherName": father_name,
        }
        if group_id is not None:
            body["groupId"] = group_id
        if student_category is not None:
            body["studentCategory"] = student_category

        payload = self._request_json(
            "POST",
            "/api/v1/registrator/users/create",
            expected_statuses=(200,),
            json_body=body,
        )
        required = {"userId", "login", "passwordClearText"}
        if isinstance(payload, dict) and required.issubset(payload.keys()):
            return payload
        raise ImportFailure("Невалидный ответ create_user")

    def add_curator_to_group(self, group_id: int, curator_id: str) -> None:
        self._request_json(
            "POST",
            f"/api/v1/groups/{group_id}/curators/{curator_id}",
            expected_statuses=(200,),
        )

    def delete_user(self, user_id: str) -> None:
        self._request(
            "DELETE",
            f"/api/v1/registrator/users/{user_id}",
            expected_statuses=(200, 204),
        )

    def delete_group(self, group_id: int) -> None:
        self._request(
            "DELETE",
            f"/api/v1/groups/{group_id}",
            expected_statuses=(200, 204),
        )


def build_credential_rows(parsed: ParsedWorkbook) -> Tuple[Dict[CuratorKey, CuratorCredentialRow], List[StudentCredentialRow]]:
    curator_groups: Dict[CuratorKey, set[str]] = defaultdict(set)
    for student in parsed.valid_students:
        curator_groups[student.curator].add(student.group_name)

    curator_rows: Dict[CuratorKey, CuratorCredentialRow] = {}
    for curator_key, groups in curator_groups.items():
        curator_rows[curator_key] = CuratorCredentialRow(
            curator=curator_key,
            groups=sorted(groups),
            status="PLANNED",
        )

    student_rows = [
        StudentCredentialRow(
            sheet=student.sheet,
            row=student.row,
            full_name=student.full_name,
            group_name=student.group_name,
            category=student.category,
            curator_full_name=student.curator.full_name,
            status="PLANNED",
        )
        for student in parsed.valid_students
    ]

    return curator_rows, student_rows


def check_clean_load(client: ApiClient) -> PreflightBaseline:
    groups = client.list_groups()
    curators = client.list_users(CURATOR_ROLE)
    students = client.list_users(STUDENT_ROLE)
    me = client.get_me()

    current_user_id = str(me.get("userId")) if me.get("userId") is not None else None

    def to_optional_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        raw = str(value).strip()
        if not raw:
            return None
        try:
            return int(raw)
        except ValueError:
            return None

    current_user_group_id = to_optional_int(me.get("groupId"))

    raw_counts = {
        "existing_groups": len(groups),
        "existing_curators": len(curators),
        "existing_students": len(students),
    }

    filtered_curators = [u for u in curators if str(u.get("userId")) != current_user_id]
    filtered_students = [u for u in students if str(u.get("userId")) != current_user_id]
    if current_user_group_id is None:
        filtered_groups = list(groups)
    else:
        filtered_groups = [g for g in groups if to_optional_int(g.get("id")) != current_user_group_id]

    effective_counts = {
        "existing_groups": len(filtered_groups),
        "existing_curators": len(filtered_curators),
        "existing_students": len(filtered_students),
    }

    ok = not (
        effective_counts["existing_groups"]
        or effective_counts["existing_curators"]
        or effective_counts["existing_students"]
    )
    message = None
    if not ok:
        message = (
            "Нарушено правило чистой загрузки (с учетом baseline-admin): "
            f"groups={effective_counts['existing_groups']}, "
            f"curators={effective_counts['existing_curators']}, "
            f"students={effective_counts['existing_students']}"
        )

    return PreflightBaseline(
        ok=ok,
        message=message,
        ignored_current_admin=current_user_id is not None or current_user_group_id is not None,
        current_user_id=current_user_id,
        current_user_group_id=current_user_group_id,
        raw_counts=raw_counts,
        effective_counts=effective_counts,
    )


def update_row_values_in_sheet(ws: Worksheet, row_idx: int, header: Dict[str, int], values: Dict[str, str]) -> None:
    mapping = {
        "student_fio": "student_fio",
        "group": "group",
        "category": "category",
        "breakfast": "breakfast",
        "lunch": "lunch",
        "comment": "comment",
        "curator": "curator",
    }
    for key, header_key in mapping.items():
        col_idx = header.get(header_key)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=values.get(key, ""))


def create_fixed_xlsx_path(source_xlsx: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return source_xlsx.with_name(f"{source_xlsx.stem}_fixed_{timestamp}{source_xlsx.suffix}")


def interactive_fix_rejects(
    *,
    xlsx_path: Path,
    rejects: Sequence[RejectRow],
    prompter: Optional[InteractivePrompter] = None,
) -> InteractiveFixResult:
    if not rejects:
        return InteractiveFixResult(
            enabled=True,
            total_issues=0,
            fixed_rows=0,
            skipped_rows=0,
            aborted=False,
            fixed_xlsx_path=None,
        )

    active_prompter = prompter or TerminalPrompter()
    wb = load_workbook(filename=xlsx_path)
    sheet_headers: Dict[str, Dict[str, int]] = {
        ws.title: parse_header([cell.value for cell in ws[1]]) for ws in wb.worksheets if ws.max_row >= 2
    }

    fixed_rows = 0
    skipped_rows = 0
    aborted = False

    for reject in sorted(rejects, key=lambda x: (x.sheet, x.row)):
        ws = wb[reject.sheet]
        header = sheet_headers.get(reject.sheet)
        if header is None:
            skipped_rows += 1
            continue

        while True:
            current_values = read_row_values_from_sheet(ws, reject.row, header)
            action = active_prompter.choose_action(reject, current_values)
            if action == "skip":
                skipped_rows += 1
                break
            if action == "abort":
                aborted = True
                break

            edited_values = dict(current_values)
            edited_values["student_fio"] = active_prompter.ask_text("ФИО студента", current_values.get("student_fio", ""))
            edited_values["group"] = active_prompter.ask_text("Группа", current_values.get("group", ""))
            edited_values["category"] = active_prompter.ask_category(current_values.get("category", ""))
            edited_values["curator"] = active_prompter.ask_text("Классный руководитель", current_values.get("curator", ""))
            edited_values["breakfast"] = active_prompter.ask_text("Завтрак", current_values.get("breakfast", ""))
            edited_values["lunch"] = active_prompter.ask_text("Обед", current_values.get("lunch", ""))
            edited_values["comment"] = active_prompter.ask_text("Комментарий", current_values.get("comment", ""))

            update_row_values_in_sheet(ws, reject.row, header, edited_values)
            _, errors = validate_row(
                sheet=reject.sheet,
                row=reject.row,
                row_values=edited_values,
                seen_students=None,
                check_duplicates=False,
            )
            if not errors:
                fixed_rows += 1
                break

            print(
                f"Строка {reject.sheet}:{reject.row} все еще невалидна: "
                + "; ".join(msg for _, msg in errors)
            )

        if aborted:
            break

    fixed_path = create_fixed_xlsx_path(xlsx_path)
    wb.save(fixed_path)

    return InteractiveFixResult(
        enabled=True,
        total_issues=len(rejects),
        fixed_rows=fixed_rows,
        skipped_rows=skipped_rows,
        aborted=aborted,
        fixed_xlsx_path=str(fixed_path.resolve()),
    )


def rollback_created(client: ApiClient, state: ImportState) -> Tuple[bool, List[str]]:
    errors: List[str] = []

    for user_id in reversed(state.created_student_ids):
        try:
            client.delete_user(user_id)
            state.deleted_student_ids.append(user_id)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"rollback delete student {user_id}: {exc}")

    for group_id in reversed(state.created_group_ids):
        try:
            client.delete_group(group_id)
            state.deleted_group_ids.append(group_id)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"rollback delete group {group_id}: {exc}")

    for user_id in reversed(state.created_curator_ids):
        try:
            client.delete_user(user_id)
            state.deleted_curator_ids.append(user_id)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"rollback delete curator {user_id}: {exc}")

    return len(errors) == 0, errors


def run_import(
    *,
    parsed: ParsedWorkbook,
    client: ApiClient,
    apply_changes: bool,
    rollback_enabled: bool,
    preflight_ok: bool = True,
) -> ImportResult:
    state = ImportState()
    group_id_map: Dict[str, int] = {}
    curator_rows, student_rows = build_credential_rows(parsed)

    rollback_performed = False
    rollback_success = False
    rollback_errors: List[str] = []

    if not apply_changes:
        for row in curator_rows.values():
            row.status = "DRY_RUN"
        for row in student_rows:
            row.status = "DRY_RUN"
        return ImportResult(
            success=True,
            preflight_ok=preflight_ok,
            applied=False,
            error=None,
            rollback_performed=False,
            rollback_success=False,
            rollback_errors=[],
            group_id_map=group_id_map,
            curator_rows=curator_rows,
            student_rows=student_rows,
            state=state,
        )

    student_row_lookup: Dict[Tuple[str, int], StudentCredentialRow] = {
        (row.sheet, row.row): row for row in student_rows
    }

    try:
        for group_name in parsed.groups:
            group_response = client.create_group(group_name)
            group_id = int(group_response["id"])
            group_id_map[group_name] = group_id
            state.created_group_ids.append(group_id)

        for curator_key in parsed.curators:
            create_response = client.create_user(
                role=CURATOR_ROLE,
                name=curator_key.name,
                surname=curator_key.surname,
                father_name=curator_key.father_name,
            )
            user_id = str(create_response["userId"])
            state.created_curator_ids.append(user_id)

            row = curator_rows[curator_key]
            row.login = str(create_response["login"])
            row.password = str(create_response["passwordClearText"])
            row.user_id = user_id
            row.status = "CREATED"

        for group_name, curators in parsed.group_curators.items():
            group_id = group_id_map[group_name]
            for curator_key in curators:
                curator_id = curator_rows[curator_key].user_id
                client.add_curator_to_group(group_id, curator_id)

        for student in parsed.valid_students:
            group_id = group_id_map[student.group_name]
            create_response = client.create_user(
                role=STUDENT_ROLE,
                name=student.name,
                surname=student.surname,
                father_name=student.father_name,
                group_id=group_id,
                student_category=student.category,
            )

            user_id = str(create_response["userId"])
            state.created_student_ids.append(user_id)

            cred_row = student_row_lookup[(student.sheet, student.row)]
            cred_row.login = str(create_response["login"])
            cred_row.password = str(create_response["passwordClearText"])
            cred_row.user_id = user_id
            cred_row.status = "CREATED"

        for row in curator_rows.values():
            if row.status == "PLANNED":
                row.status = "NOT_CREATED"
        for row in student_rows:
            if row.status == "PLANNED":
                row.status = "NOT_CREATED"

        return ImportResult(
            success=True,
            preflight_ok=preflight_ok,
            applied=True,
            error=None,
            rollback_performed=False,
            rollback_success=False,
            rollback_errors=[],
            group_id_map=group_id_map,
            curator_rows=curator_rows,
            student_rows=student_rows,
            state=state,
        )

    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)

        if rollback_enabled:
            rollback_performed = True
            rollback_success, rollback_errors = rollback_created(client, state)

        deleted_curators = set(state.deleted_curator_ids)
        deleted_students = set(state.deleted_student_ids)

        for row in curator_rows.values():
            if row.status == "CREATED":
                if row.user_id in deleted_curators:
                    row.status = "ROLLED_BACK"
                elif rollback_performed:
                    row.status = "CREATED_ROLLBACK_FAILED"
                else:
                    row.status = "CREATED"
            elif row.status == "PLANNED":
                row.status = "NOT_CREATED"

        for row in student_rows:
            if row.status == "CREATED":
                if row.user_id in deleted_students:
                    row.status = "ROLLED_BACK"
                elif rollback_performed:
                    row.status = "CREATED_ROLLBACK_FAILED"
                else:
                    row.status = "CREATED"
            elif row.status == "PLANNED":
                row.status = "NOT_CREATED"

        return ImportResult(
            success=False,
            preflight_ok=preflight_ok,
            applied=True,
            error=error_text,
            rollback_performed=rollback_performed,
            rollback_success=rollback_success,
            rollback_errors=rollback_errors,
            group_id_map=group_id_map,
            curator_rows=curator_rows,
            student_rows=student_rows,
            state=state,
        )


def write_reject_csv(path: Path, rejects: Sequence[RejectRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(REJECT_HEADER)
        for row in rejects:
            writer.writerow(
                [
                    row.sheet,
                    row.row,
                    row.error_code,
                    row.error_message,
                    row.raw_fio,
                    row.raw_group,
                    row.raw_category,
                    row.raw_curator,
                ]
            )


def write_credentials_xlsx(
    path: Path,
    curator_rows: Dict[CuratorKey, CuratorCredentialRow],
    student_rows: Sequence[StudentCredentialRow],
) -> None:
    wb = Workbook()

    ws_curators = wb.active
    ws_curators.title = "Кураторы"
    ws_curators.append(CURATORS_SHEET_HEADER)

    for row in sorted(curator_rows.values(), key=lambda x: (x.curator.surname, x.curator.name, x.curator.father_name)):
        ws_curators.append(
            [
                row.curator.full_name,
                ", ".join(sorted(row.groups)),
                row.login,
                row.password,
                row.user_id,
                row.status,
            ]
        )

    ws_students = wb.create_sheet("Студенты")
    ws_students.append(STUDENTS_SHEET_HEADER)

    for row in sorted(student_rows, key=lambda x: (x.sheet, x.row)):
        ws_students.append(
            [
                row.sheet,
                row.row,
                row.full_name,
                row.group_name,
                row.category,
                row.curator_full_name,
                row.login,
                row.password,
                row.user_id,
                row.status,
            ]
        )

    wb.save(path)

    if os.name != "nt":
        os.chmod(path, 0o600)


def build_report(
    *,
    args: argparse.Namespace,
    parsed: ParsedWorkbook,
    result: ImportResult,
    preflight: PreflightBaseline,
    interactive_fix: InteractiveFixResult,
    started_at: datetime,
    finished_at: datetime,
) -> Dict[str, Any]:
    duration_seconds = round((finished_at - started_at).total_seconds(), 3)

    report = {
        "mode": "apply" if args.apply else "dry-run",
        "input": {
            "xlsx": str(Path(args.xlsx).resolve()),
            "baseUrl": args.base_url,
            "sheets": parsed.sheets_seen,
        },
        "timing": {
            "startedAt": started_at.isoformat(),
            "finishedAt": finished_at.isoformat(),
            "durationSeconds": duration_seconds,
        },
        "counts": {
            "processedStudentRows": parsed.processed_student_rows,
            "skippedNonStudentRows": parsed.skipped_non_student_rows,
            "validRows": len(parsed.valid_students),
            "rejectRows": len(parsed.rejects),
            "plannedGroups": len(parsed.groups),
            "plannedCurators": len(parsed.curators),
            "plannedStudents": len(parsed.valid_students),
            "createdGroups": len(result.state.created_group_ids),
            "createdCurators": len(result.state.created_curator_ids),
            "createdStudents": len(result.state.created_student_ids),
            "deletedGroupsRollback": len(result.state.deleted_group_ids),
            "deletedCuratorsRollback": len(result.state.deleted_curator_ids),
            "deletedStudentsRollback": len(result.state.deleted_student_ids),
        },
        "quality": {
            "breakfastValues": dict(parsed.breakfast_values),
            "lunchValues": dict(parsed.lunch_values),
            "commentsNonEmpty": parsed.comments_non_empty,
        },
        "preflight": {
            "ok": preflight.ok,
            "message": preflight.message,
            "existingGroups": preflight.effective_counts.get("existing_groups", -1),
            "existingCurators": preflight.effective_counts.get("existing_curators", -1),
            "existingStudents": preflight.effective_counts.get("existing_students", -1),
        },
        "preflightBaseline": {
            "currentUserId": preflight.current_user_id,
            "currentUserGroupId": preflight.current_user_group_id,
            "ignoredCurrentAdmin": preflight.ignored_current_admin,
            "rawCounts": preflight.raw_counts,
            "effectiveCounts": preflight.effective_counts,
        },
        "interactiveFix": {
            "enabled": interactive_fix.enabled,
            "totalIssues": interactive_fix.total_issues,
            "fixedRows": interactive_fix.fixed_rows,
            "skippedRows": interactive_fix.skipped_rows,
            "aborted": interactive_fix.aborted,
            "fixedXlsxPath": interactive_fix.fixed_xlsx_path,
        },
        "result": {
            "success": result.success,
            "applied": result.applied,
            "error": result.error,
            "rollbackPerformed": result.rollback_performed,
            "rollbackSuccess": result.rollback_success,
            "rollbackErrors": result.rollback_errors,
        },
        "artifacts": {
            "rejectCsv": str(Path(args.out_dir, "reject.csv").resolve()),
            "credentialsXlsx": str(Path(args.out_dir, "credentials.xlsx").resolve()),
            "reportJson": str(Path(args.out_dir, "report.json").resolve()),
            "fixedXlsx": interactive_fix.fixed_xlsx_path,
        },
    }
    return report


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Импорт групп/кураторов/студентов из XLSX через API")
    parser.add_argument("--xlsx", required=True, help="Путь к входному .xlsx")
    parser.add_argument("--base-url", required=True, help="Базовый URL backend API, например http://localhost:8080")
    parser.add_argument("--login", required=True, help="Логин администратора/регистратора")
    parser.add_argument("--password-env", required=True, help="Имя переменной окружения с паролем")
    parser.add_argument("--out-dir", required=True, help="Папка для report.json/reject.csv/credentials.xlsx")
    parser.add_argument("--apply", action="store_true", help="Выполнить запись в API. Без флага работает dry-run")
    parser.add_argument(
        "--disable-rollback",
        action="store_true",
        help="Отключить rollback при ошибке во время apply",
    )
    parser.add_argument("--timeout", type=float, default=15.0, help="HTTP timeout в секундах")
    parser.add_argument(
        "--max-attempts",
        type=int,
        default=4,
        help="Максимум попыток для network/5xx запросов",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    password = os.environ.get(args.password_env)
    if not password:
        raise ImportFailure(f"Переменная окружения '{args.password_env}' не задана")

    started_at = datetime.now(timezone.utc)
    source_xlsx = Path(args.xlsx)
    parsed = parse_workbook(source_xlsx)

    interactive_fix = InteractiveFixResult(
        enabled=True,
        total_issues=0,
        fixed_rows=0,
        skipped_rows=0,
        aborted=False,
        fixed_xlsx_path=None,
    )

    if parsed.rejects:
        interactive_fix = interactive_fix_rejects(
            xlsx_path=source_xlsx,
            rejects=parsed.rejects,
        )
        if interactive_fix.fixed_xlsx_path:
            parsed = parse_workbook(Path(interactive_fix.fixed_xlsx_path))

    preflight = PreflightBaseline(
        ok=False,
        message="Preflight не выполнялся",
        ignored_current_admin=False,
        current_user_id=None,
        current_user_group_id=None,
        raw_counts={"existing_groups": -1, "existing_curators": -1, "existing_students": -1},
        effective_counts={"existing_groups": -1, "existing_curators": -1, "existing_students": -1},
    )

    if interactive_fix.aborted:
        curator_rows, student_rows = build_credential_rows(parsed)
        for row in curator_rows.values():
            row.status = "ABORTED_BY_USER"
        for row in student_rows:
            row.status = "ABORTED_BY_USER"
        result = ImportResult(
            success=False,
            preflight_ok=False,
            applied=False,
            error="Интерактивная правка прервана пользователем",
            rollback_performed=False,
            rollback_success=False,
            rollback_errors=[],
            group_id_map={},
            curator_rows=curator_rows,
            student_rows=student_rows,
            state=ImportState(),
        )
    else:
        client = ApiClient(
            base_url=args.base_url,
            timeout_seconds=args.timeout,
            max_attempts=args.max_attempts,
        )

        preflight_error: Optional[str] = None
        try:
            client.login(args.login, password)
            preflight = check_clean_load(client)
            if not preflight.ok:
                preflight_error = preflight.message or "Preflight не пройден"
        except Exception as exc:  # noqa: BLE001
            preflight_error = str(exc)
            preflight = PreflightBaseline(
                ok=False,
                message=preflight_error,
                ignored_current_admin=False,
                current_user_id=None,
                current_user_group_id=None,
                raw_counts={"existing_groups": -1, "existing_curators": -1, "existing_students": -1},
                effective_counts={"existing_groups": -1, "existing_curators": -1, "existing_students": -1},
            )

        if preflight_error is not None:
            curator_rows, student_rows = build_credential_rows(parsed)
            for row in curator_rows.values():
                row.status = "PRECHECK_FAILED"
            for row in student_rows:
                row.status = "PRECHECK_FAILED"

            result = ImportResult(
                success=False,
                preflight_ok=False,
                applied=False,
                error=preflight_error,
                rollback_performed=False,
                rollback_success=False,
                rollback_errors=[],
                group_id_map={},
                curator_rows=curator_rows,
                student_rows=student_rows,
                state=ImportState(),
            )
        else:
            result = run_import(
                parsed=parsed,
                client=client,
                apply_changes=bool(args.apply),
                rollback_enabled=not args.disable_rollback,
                preflight_ok=preflight.ok,
            )

    reject_path = out_dir / "reject.csv"
    credentials_path = out_dir / "credentials.xlsx"
    report_path = out_dir / "report.json"

    write_reject_csv(reject_path, parsed.rejects)
    write_credentials_xlsx(credentials_path, result.curator_rows, result.student_rows)

    finished_at = datetime.now(timezone.utc)

    report = build_report(
        args=args,
        parsed=parsed,
        result=result,
        preflight=preflight,
        interactive_fix=interactive_fix,
        started_at=started_at,
        finished_at=finished_at,
    )

    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "success": result.success,
                "mode": "apply" if args.apply else "dry-run",
                "processed": parsed.processed_student_rows,
                "valid": len(parsed.valid_students),
                "reject": len(parsed.rejects),
                "groups": len(parsed.groups),
                "curators": len(parsed.curators),
                "students": len(parsed.valid_students),
                "interactiveAborted": interactive_fix.aborted,
                "fixedXlsx": interactive_fix.fixed_xlsx_path,
                "outDir": str(out_dir.resolve()),
            },
            ensure_ascii=False,
        )
    )

    return 0 if result.success else 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportFailure as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)
    except KeyboardInterrupt:
        print("ERROR: interrupted", file=sys.stderr)
        raise SystemExit(130)
